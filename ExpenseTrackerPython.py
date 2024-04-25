import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
import os
import re
import pyodbc
import smtplib
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import time
import babel.numbers

MileageTemplatePath = r"Q:\A Portocarrero\Mileage Report Template.xlsx"

# Define the reimbursement rate per mile
REIMBURSEMENT_RATE = 0.545

# Define the distances between locations for round trip
DISTANCES = {
    ("CCC", "NLR"): 25,
    ("CCC", "Conway"): 60,
    ("CCC", "El Dorado"): 246,
    ("CCC", "Pine Bluff"): 100,
    ("CCC", "Searcy"): 110,
    ("CCC", "Baptist"): 3,
    ('Searcy', 'NLR'): 88, 
    ('Searcy', 'Pine Bluff'): 184, 
    ('Searcy', 'Conway'): 104,
    ('Searcy', 'El Dorado'): 330,
    ('Searcy', 'Baptist'): 112,
    ('NLR', 'Pine Bluff'): 98, 
    ('NLR', 'Conway'): 63, 
    ('NLR', 'El Dorado'): 244,
    ('NLR', 'Baptist'): 27, 
    ('Pine Bluff', 'Conway'): 150, 
    ('Pine Bluff', 'El Dorado'): 170, 
    ('Pine Bluff', 'Baptist'): 102, 
    ('Conway', 'El Dorado'): 294, 
    ('Conway', 'Baptist'): 62,
    ('El Dorado', 'Baptist'): 248
}

# Initialize last email time
last_email_time = 0
current_row = 12  # Starting row for the first entry

def calculate_reimbursement(site_origin, site_destination, round_trip):
    # Check if the distance exists for the provided sites
    distance = DISTANCES.get((site_origin, site_destination))
    if distance is None:
        # If not, check the reversed order
        distance = DISTANCES.get((site_destination, site_origin))

    if distance is not None:
        if round_trip:  # If it's a round trip, use the full distance
            reimbursement_amount = distance * REIMBURSEMENT_RATE
        else:  # If it's one way, divide the distance by 2
            reimbursement_amount = (distance / 2) * REIMBURSEMENT_RATE
        return distance, reimbursement_amount
    else:
        return None, None

def update_reimbursement_message():
    origin_site = origin_site_combobox.get()  # Fetch selected origin site
    site = site_combobox.get()

    distance, reimbursement_amount = calculate_reimbursement(origin_site, site, round_trip_checkbox_var.get())
    if distance is not None and reimbursement_amount is not None:
        reimbursement_amount_var.set(f"Total Reimbursement Amount: ${reimbursement_amount:.2f} USD")
    else:
        reimbursement_amount_var.set("Distance information not available for selected sites.")

def site_combobox_changed(event):
    update_reimbursement_message()

def round_trip_checkbox_changed():
    update_reimbursement_message()

def submit_button_click():
    global last_email_time, current_row
    name = name_combobox.get()
    date = date_entry.get()
    # Extract numeric value from the label text and remove non-numeric characters
    amount_text = reimbursement_amount_var.get().replace("Total Reimbursement Amount: $", "").strip()
    amount_text = re.sub(r'\D', '', amount_text)  # Remove non-numeric characters
    # Convert the numeric value to float
    try:
        amount = float(amount_text) / 100  # Divide by 100 to convert cents to dollars
    except ValueError:
        messagebox.showerror("Error Submitting Entries", "Invalid amount format. Please enter a valid number.")
        return
    site = site_combobox.get()

    # Update the report spreadsheet
    try:
        origin_site = origin_site_combobox.get()  # Fetch origin site
        distance, reimbursement_amount = calculate_reimbursement(origin_site, site, round_trip_checkbox_var.get())
        if distance is not None and reimbursement_amount is not None:
            update_report_spreadsheet(date, origin_site, site, reimbursement_amount, distance, current_row)
            current_row += 1  # Move to the next row for the next entry
    except Exception as e:
        print("Error updating report spreadsheet:", e)
        messagebox.showerror("Error Submitting Entries", "Error updating report spreadsheet. Excel Mileage Template Might be Open")
        return

    # If an error occurred during report update, stop further actions
    try:
        # Insert data into the Access database
        conn = pyodbc.connect(r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=Q:\A Portocarrero\Database11.accdb;Trusted_Connection=yes;')
        cursor = conn.cursor()

        cursor.execute("INSERT INTO Table1 (DateIncurrence, Amount, Site, Account, IncurredBy, Notes) VALUES (?, ?, ?, ?, ?, ?)", 
                        (date, amount, site, "Travel Expenses", name, "Needs Approval"))
        
        conn.commit()
        conn.close()
        print("Entry successfully added to the database.")
        
        # Send email notification if an hour has passed since the last email
        current_time = time.time()
        if current_time - last_email_time >= 3600:
            send_email_notification(name, date, amount, site)
            last_email_time = current_time
    except Exception as e:
        print("Error:", e)
        messagebox.showerror("Error Submitting Entries", "May not have access to the Database")

def send_email_notification(name, date, amount, site):
    sender_email = "EntryMessageCarti@outlook.com"
    receiver_email = "andres.bonifaz@carti.com"
    password = "physics123"

    message = MIMEMultipart("alternative")
    message["Subject"] = "New Travel Entry Submitted"
    message["From"] = sender_email
    message["To"] = receiver_email

    text = f"""\
    A new travel entry has been submitted.
    Name: {name}
    """
    html = f"""\
    <html>
      <body>
        <p>A new travel entry has been submitted.</p>
        <p><strong>Name:</strong> {name}</p>
      </body>
    </html>
    """

    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")

    message.attach(part1)
    message.attach(part2)

    with smtplib.SMTP('smtp.office365.com', 587) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())
        print('Email sent!')

def update_report_spreadsheet(date, origin_site, destination_site, reimbursement_amount, distance, current_row):
    # Define a dictionary to map site names to addresses
    site_addresses = {
        "CCC": "8901 CARTI Way, LR",
        "Baptist": "9601 Bapt Health Dr, LR",
        "NLR": "3401 Springhill Drive, NLR",
        "Conway": "2605 College Ave., Conway",
        "El Dorado": "1601 North West Avenue, El Dorado",
        "Pine Bluff": "5001 Bobo Road, Pine Bluff",
        "Searcy": "405 Rodgers Drive, Searcy"
    }

    # Open the report spreadsheet
    report_path = MileageTemplatePath
    wb = load_workbook(report_path)
    sheet = wb.active

    # Calculate the total distance based on whether it's a round trip or not
    if round_trip_checkbox_var.get():
        total_distance = distance
    else:
        total_distance = distance / 2
    
    origin_address = site_addresses.get(origin_site, "Unknown Site")
    destination_address = site_addresses.get(destination_site, "Unknown Site")
    
    sheet.cell(row=current_row, column=1, value=date)
    sheet.cell(row=current_row, column=2, value=origin_address)
    sheet.cell(row=current_row, column=4, value=destination_address)
    sheet.cell(row=current_row, column=9, value=total_distance)  # Use total_distance here
    sheet.cell(row=current_row, column=10, value=reimbursement_amount)

    # Save the changes
    wb.save(report_path)


def generate_report():
    # Open the report template
    report_path = MileageTemplatePath
    try:
        wb = load_workbook(filename=report_path)
        sheet = wb.active
        
        # Replace cell A3 with "NAME: [user]"
        user = name_combobox.get()
        sheet['A3'] = f"NAME: {user}"
        sheet['A3'].font = Font(bold=True)
        
        # Replace cell F54 with the current date
        today = datetime.datetime.today().strftime('%m-%d-%Y')
        sheet['F54'] = today
        
        # Calculate the sum of amounts from J12 to J29
        total_amount = sum(float(sheet.cell(row=i, column=10).value) for i in range(12, 30) if sheet.cell(row=i, column=10).value is not None)
        # Write the sum to cell J50
        sheet['J50'] = total_amount
        
        # Save the changes
        wb.save(report_path)

        # Open the updated report spreadsheet
        os.startfile(report_path)
    except Exception as e:
        print("Error:", e)
        messagebox.showerror("Error Generating Report", "1) Excel Mileage Template Might be Open \n2) You didn't submit any entries")

def clear_report_spreadsheet():
    # Open the report spreadsheet
    report_path = MileageTemplatePath
    wb = openpyxl.load_workbook(report_path)
    sheet = wb.active

    # Clear the content of specified columns for rows 12 to 29
    columns_to_clear = [1, 2, 4, 9, 10]
    for row in range(12, 30):
        for col in columns_to_clear:
            sheet.cell(row=row, column=col).value = None

    # Save the changes
    wb.save(report_path)

def update_travel_site_options(selected_origin_site):
    # Copy the original sites list
    available_sites = sites.copy()

    # Remove the selected origin site from the available sites list
    if selected_origin_site:
        available_sites.remove(selected_origin_site)

    # Update the options for the travel site combobox
    site_combobox['values'] = available_sites

def origin_site_combobox_changed(event):
    update_reimbursement_message()
    selected_origin_site = origin_site_combobox.get()
    update_travel_site_options(selected_origin_site=selected_origin_site)

root = tk.Tk()
root.title("Travel Entry Form")

main_frame = ttk.Frame(root, padding="20")
main_frame.grid(row=0, column=0)

name_label = ttk.Label(main_frame, text="Name:")
name_label.grid(row=0, column=0, sticky="w")

users = [
    "Scott", "Paul", "Joe", "Vance", "Cathy Large",  "Regan Hime", "Ashley Matsushita", "Erika Brescia", 
    "Andres", "Rob Moller", "Tyler", "Michael Johnson", "Misty Tholl", "Dianna", "Scott Legg",  
    "Doug", "Amy", "Bob", "Megan", "Ashley Hicks", "Micheal Hall"
]
name_combobox = ttk.Combobox(main_frame, values=users, state="readonly")
name_combobox.grid(row=0, column=1)

date_label = ttk.Label(main_frame, text="Date of Travel:")
date_label.grid(row=1, column=0, sticky="w")

# Use DateEntry widget for interactive calendar
date_entry = DateEntry(main_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
date_entry.grid(row=1, column=1, sticky="ew")

# Define available sites
sites = [
    "CCC", "Searcy", "NLR", "Pine Bluff", "Conway", "El Dorado", "Baptist"
]

# Create a combobox for selecting the origin site
origin_site_label = ttk.Label(main_frame, text="Site of Origin:")
origin_site_label.grid(row=2, column=0, sticky="w")

origin_site_combobox = ttk.Combobox(main_frame, values=sites, state="readonly")
origin_site_combobox.bind("<<ComboboxSelected>>", origin_site_combobox_changed)
origin_site_combobox.grid(row=2, column=1)
origin_site_combobox.current(0)  # Set default value to "CCC"


# Create a combobox for selecting the destination site
site_label = ttk.Label(main_frame, text="Site of Travel:")
site_label.grid(row=3, column=0, sticky="w")

site_combobox = ttk.Combobox(main_frame, values=sites, state="readonly")
site_combobox.grid(row=3, column=1)
site_combobox.bind("<<ComboboxSelected>>", site_combobox_changed)

# Initialize the travel site combobox options based on the default origin site
default_origin_site = origin_site_combobox.get()
update_travel_site_options(default_origin_site)

# Add a checkbox for Round Trip
round_trip_checkbox_var = tk.BooleanVar()
round_trip_checkbox = ttk.Checkbutton(main_frame, variable=round_trip_checkbox_var, text="Round Trip", command=round_trip_checkbox_changed)
round_trip_checkbox.grid(row=3, column=2, padx=(10, 0), pady=5, sticky="w")

reimbursement_amount_var = tk.StringVar()  # Use StringVar for label text
reimbursement_amount_label = ttk.Label(main_frame, textvariable=reimbursement_amount_var)
reimbursement_amount_label.grid(row=4, column=0, columnspan=3, pady=5)

submit_button = ttk.Button(main_frame, text="Submit Entry", command=submit_button_click)
submit_button.grid(row=5, column=0, columnspan=3, pady=10)

generate_report_button = ttk.Button(main_frame, text="Generate Report", command=generate_report)
generate_report_button.grid(row=6, column=0, columnspan=3, pady=10)

for child in main_frame.winfo_children():
    child.grid_configure(padx=5, pady=5)

try:
    clear_report_spreadsheet()
except Exception as e:
    print("Error clearing spreadsheet:", e)
    messagebox.showerror("Error Clearing Spreadsheet", "Close the Mileage Template")
    root.destroy()

root.mainloop()

